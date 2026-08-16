"""Parse SNOW T15/T23 xlsx into CorpusRecord triples (原文 ↔ やさしい日本語 ↔ English).

REAL SCHEMA — inspected in the actual 2020-01-07 release files, 2026-08-07.
Do not trust the lab pages alone; the files are the authority.

T15 (``T15-2020.1.7.xlsx``): one sheet ``Sheet1``, dims A1:D50001 →
50,000 data rows. Header::

    ID | #日本語(原文) | #やさしい日本語 | #英語(原文)

``ID`` is an int 1..50000 (unique, contiguous). All three text columns are
non-empty strings in every row. English is lowercased/tokenised MT-style text
(from small_parallel_enja). No proper-noun column, no 【】 markers anywhere.

T23 (``T23-2020.1.7.xlsx``): TWO sheets.

1. ``平易化コーパス`` — dims A1:E34301 → 34,300 data rows. Header::

       ID | #日本語(原文) | #やさしい日本語 | #英語(原文) | #固有名詞

   ``ID`` is ``<worker>_<n>`` (7 workers Ab/Ac/Ae/Af/Ah/Ak/Al × 4,900 each).
   T23's proper-noun convention is this FIFTH COLUMN, not inline markup:
   workers extracted proper nouns they were told not to rewrite
   (``固有名詞は書き換えないように指示をしており、固有名詞の判断は作業者に
   任せています`` — lab page). Populated in 2,812 of 34,300 rows.
   【】 appears in exactly ONE row (``Af_2708``), where the sentence itself
   discusses the suffix 【-ion】 — sentence content, not a corpus convention.
   One cell quirk: ``Al_2776``'s 固有名詞 was typed as a clock time and Excel
   stored it as a time value; openpyxl yields ``datetime.time(12, 0)``. We
   serialise such cells with ``.isoformat()`` ("12:00:00").

2. ``評価用`` — dims A1:Q101 → 100 data rows, WIDE format: the same 100
   originals simplified independently by all 7 workers. Header::

       ID | #日本語(原文) | Ab | Ac | Af | Ae | Ah | Ak | Al | #英語(原文)
          | Ab_固有名詞 | ... | Al_固有名詞

   ``ID`` is ``eval_1``..``eval_100``. We emit one record per (row, worker):
   100 × 7 = 700 records — which is exactly how the lab reaches its stated
   35,000 (34,300 + 700). Different workers often produce IDENTICAL
   simplifications of the same original, so the eval sheet is the dominant
   source of exact-duplicate triples; duplicates are kept and flagged, never
   dropped (see ``flag_exact_duplicates``).

Record shape (one record per original↔simplified↔english triple):

- id: ``snow-t15:<ID>`` / ``snow-t23:<ID>`` / ``snow-t23:<ID>:<worker>``
  (eval-sheet records carry the worker code; all ids verified unique).
- source: ``snow-t15`` | ``snow-t23``; title: ``""`` (sentences have none).
- text: the original Japanese (#日本語(原文)).
- meta: ``corpus`` ("t15"|"t23"), ``simplified``, ``english``; T23 adds
  ``sheet`` (the real sheet name, 平易化コーパス or 評価用), eval rows add
  ``worker``, and ``proper_nouns`` appears only when the cell is non-empty.
  Flagged duplicates add ``exact_duplicate_of: <id of first occurrence>``.

Normalisation — the ONLY cleaning applied: leading/trailing whitespace strip
per cell. In the real data this touches 8 T15 cells (7 English cells with a
leading space from quote fragments, e.g. ID 40537 ``" hello , people of the
world ! ''"``; and ID 46615's やさしい日本語 which ends in ``\\n``) and the
time-cell serialisation above. Internal whitespace, casing, punctuation and
the MT tokenisation of the English are preserved byte-for-byte.

Rows whose cells are all empty are skipped (xlsx phantom rows); rows with a
non-empty cell but an empty original are counted in stats as
``skipped_no_original`` and skipped, because CorpusRecord requires non-empty
text. Both counts are 0 in the real files (asserted in the realdata tier).
"""

from __future__ import annotations

import datetime as dt
import json
import sys
from dataclasses import replace
from pathlib import Path
from typing import Any, Iterable, Iterator

import openpyxl

from corpus.records import CorpusRecord, write_jsonl

T15_SHEET = "Sheet1"
T15_HEADER = ("ID", "#日本語(原文)", "#やさしい日本語", "#英語(原文)")
T23_MAIN_SHEET = "平易化コーパス"
T23_MAIN_HEADER = ("ID", "#日本語(原文)", "#やさしい日本語", "#英語(原文)", "#固有名詞")
T23_EVAL_SHEET = "評価用"


class ParseError(ValueError):
    pass


def _cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (dt.datetime, dt.date, dt.time)):
        return v.isoformat()
    return str(v).strip()


def _rows(path: str | Path, sheet: str) -> Iterator[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ParseError(f"{path}: sheet {sheet!r} not found (has {wb.sheetnames})")
        yield from wb[sheet].iter_rows(values_only=True)
    finally:
        wb.close()


def _check_header(got: tuple, want: tuple, origin: str) -> None:
    if tuple(_cell(v) for v in got[: len(want)]) != want:
        raise ParseError(f"{origin}: header drift — expected {want}, got {got}")


def parse_t15(path: str | Path, stats: dict | None = None) -> Iterator[CorpusRecord]:
    rows = _rows(path, T15_SHEET)
    _check_header(next(rows), T15_HEADER, f"{path}!{T15_SHEET}")
    for row in rows:
        cells = [_cell(v) for v in row[:4]]
        if not any(cells):
            continue
        rid, original, simplified, english = cells
        if not original:
            if stats is not None:
                stats["skipped_no_original"] = stats.get("skipped_no_original", 0) + 1
            continue
        yield CorpusRecord(
            id=f"snow-t15:{rid}",
            source="snow-t15",
            title="",
            text=original,
            meta={"corpus": "t15", "simplified": simplified, "english": english},
        )


def _t23_main(path: str | Path, stats: dict | None) -> Iterator[CorpusRecord]:
    rows = _rows(path, T23_MAIN_SHEET)
    _check_header(next(rows), T23_MAIN_HEADER, f"{path}!{T23_MAIN_SHEET}")
    for row in rows:
        cells = [_cell(v) for v in row[:5]]
        if not any(cells):
            continue
        rid, original, simplified, english, proper_nouns = cells
        if not original:
            if stats is not None:
                stats["skipped_no_original"] = stats.get("skipped_no_original", 0) + 1
            continue
        meta: dict[str, Any] = {
            "corpus": "t23",
            "sheet": T23_MAIN_SHEET,
            "simplified": simplified,
            "english": english,
        }
        if proper_nouns:
            meta["proper_nouns"] = proper_nouns
        yield CorpusRecord(id=f"snow-t23:{rid}", source="snow-t23", title="", text=original, meta=meta)


def _t23_eval(path: str | Path, stats: dict | None) -> Iterator[CorpusRecord]:
    rows = _rows(path, T23_EVAL_SHEET)
    header = [_cell(v) for v in next(rows)]
    if header[:2] != ["ID", "#日本語(原文)"] or "#英語(原文)" not in header:
        raise ParseError(f"{path}!{T23_EVAL_SHEET}: header drift — got {header}")
    en_col = header.index("#英語(原文)")
    workers = header[2:en_col]
    if not workers or any("固有名詞" in w for w in workers):
        raise ParseError(f"{path}!{T23_EVAL_SHEET}: worker columns drift — got {header}")
    pn_cols = {w: header.index(f"{w}_固有名詞") for w in workers if f"{w}_固有名詞" in header}
    for row in rows:
        cells = [_cell(v) for v in row]
        if not any(cells):
            continue
        rid, original = cells[0], cells[1]
        if not original:
            if stats is not None:
                stats["skipped_no_original"] = stats.get("skipped_no_original", 0) + 1
            continue
        english = cells[en_col]
        for i, worker in enumerate(workers, start=2):
            meta: dict[str, Any] = {
                "corpus": "t23",
                "sheet": T23_EVAL_SHEET,
                "worker": worker,
                "simplified": cells[i],
                "english": english,
            }
            pn = cells[pn_cols[worker]] if worker in pn_cols and pn_cols[worker] < len(cells) else ""
            if pn:
                meta["proper_nouns"] = pn
            yield CorpusRecord(
                id=f"snow-t23:{rid}:{worker}", source="snow-t23", title="", text=original, meta=meta
            )


def parse_t23(path: str | Path, stats: dict | None = None) -> Iterator[CorpusRecord]:
    yield from _t23_main(path, stats)
    yield from _t23_eval(path, stats)


def flag_exact_duplicates(records: Iterable[CorpusRecord]) -> list[CorpusRecord]:
    """Flag later exact repeats of a (corpus, text, simplified, english) triple.

    Keeps every record; later occurrences gain meta["exact_duplicate_of"] =
    the id of the first occurrence. Never drops, never reorders.
    """
    first: dict[tuple, str] = {}
    out: list[CorpusRecord] = []
    for rec in records:
        key = (rec.meta.get("corpus"), rec.text, rec.meta.get("simplified"), rec.meta.get("english"))
        if key in first:
            out.append(replace(rec, meta={**rec.meta, "exact_duplicate_of": first[key]}))
        else:
            first[key] = rec.id
            out.append(rec)
    return out


def parse_all(data_dir: str | Path, stats: dict | None = None) -> list[CorpusRecord]:
    from corpus.sources.snow.fetch import FILENAMES

    data_dir = Path(data_dir)
    records = list(parse_t15(data_dir / FILENAMES["t15"], stats))
    records += list(parse_t23(data_dir / FILENAMES["t23"], stats))
    return flag_exact_duplicates(records)


def compute_stats(records: Iterable[CorpusRecord]) -> dict:
    stats: dict[str, Any] = {"total": 0, "per_corpus": {}}
    for rec in records:
        stats["total"] += 1
        c = stats["per_corpus"].setdefault(
            rec.meta.get("corpus", "?"),
            {"rows": 0, "complete_triples": 0, "exact_duplicates_flagged": 0, "with_proper_nouns": 0},
        )
        c["rows"] += 1
        if rec.text and rec.meta.get("simplified") and rec.meta.get("english"):
            c["complete_triples"] += 1
        if "exact_duplicate_of" in rec.meta:
            c["exact_duplicates_flagged"] += 1
        if "proper_nouns" in rec.meta:
            c["with_proper_nouns"] += 1
    for c in stats["per_corpus"].values():
        c["complete_triple_rate"] = round(c["complete_triples"] / c["rows"], 6) if c["rows"] else 0.0
    return stats


def main(argv: list[str] | None = None) -> int:
    args = argv if argv is not None else sys.argv[1:]
    data_dir = Path(args[0]) if args else Path("data/snow")
    out_dir = Path(args[1]) if len(args) > 1 else Path("out/snow")
    parse_stats: dict = {}
    records = parse_all(data_dir, parse_stats)
    stats = compute_stats(records) | {"skipped_no_original": parse_stats.get("skipped_no_original", 0)}
    out_dir.mkdir(parents=True, exist_ok=True)
    n = write_jsonl(records, out_dir / "snow.jsonl")
    (out_dir / "stats.json").write_text(
        json.dumps(stats, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"{n} records → {out_dir / 'snow.jsonl'}")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
