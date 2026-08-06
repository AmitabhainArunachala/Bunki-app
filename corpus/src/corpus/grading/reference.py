"""Reference distribution builder + minimal SNOW T15 loader.

The disagreement mapping in grade.py bins each signal against tercile edges
computed over per-sentence signal values from SNOW T15 original sentences.
This module builds that committed file (``reference_snow_t15.json``).

SNOW T15 (Maruyama & Yamamoto, "Simplified Corpus with Core Vocabulary",
LREC 2018; CC BY 4.0; https://www.jnlp.org/GengoHouse/snow/t15): 50,000
original↔simplified Japanese sentence pairs, distributed as one xlsx with
columns ID / #日本語(原文) / #やさしい日本語 / #英語(原文). The download URL
is served by https://www.jnlp.org/cgi-priv/download.cgi?id=SNOW/T15 via a
meta-refresh redirect to jnlp's file host.

MERGE NOTE: a parallel branch ingests SNOW properly (loader + PROVENANCE).
This loader is deliberately minimal and test/reference-only; when the SNOW
ingestion branch lands, this should collapse onto it.

Build (realdata; requires the xlsx + the NINJAL download):

    python -m corpus.grading.reference data/snow/T15-2020.1.7.xlsx
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import json
import random
import statistics
import sys
from pathlib import Path

from corpus.grading._mecab import get_tagger
from corpus.grading.jread import jread_signal
from corpus.grading.lexical import lexical_signal
from corpus.grading.tmr import tmr_signal

SNOW_URL = "https://www.jnlp.org/cgi-priv/download.cgi?id=SNOW/T15"
REFERENCE_VERSION = "v1"


def load_snow_pairs(xlsx_path: str | Path) -> list[tuple[str, str]]:
    """[(original, simplified), ...] from the SNOW T15 xlsx."""
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    pairs: list[tuple[str, str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header: ID / #日本語(原文) / #やさしい日本語 / #英語(原文)
        if row[1] and row[2]:
            pairs.append((str(row[1]).strip(), str(row[2]).strip()))
    wb.close()
    return pairs


def sample_pairs(pairs: list[tuple[str, str]], n: int, seed: int) -> list[tuple[str, str]]:
    """Uniform sample without replacement, fixed seed — stated, reproducible."""
    rng = random.Random(seed)
    return rng.sample(pairs, n)


def build_reference(
    xlsx_path: str | Path, *, n: int = 2000, seed: int = 42
) -> dict:
    from corpus.grading.ninjal import load_kokugoken

    tagger = get_tagger()
    core_level = load_kokugoken().core.substrate
    pairs = load_snow_pairs(xlsx_path)
    originals = [orig for orig, _simp in sample_pairs(pairs, n, seed)]

    jread_vals, cov_vals, tmr_vals = [], [], []
    for text in originals:
        jread_vals.append(jread_signal(text, tagger)["score"])
        lex = lexical_signal(text, tagger=tagger)
        t = tmr_signal(text, tagger=tagger)
        if lex["coverage"] is not None:
            cov_vals.append(lex["coverage"])
        if core_level in t["by_level"]:
            tmr_vals.append(t["by_level"][core_level])

    def terciles(vals: list[float]) -> list[float]:
        return statistics.quantiles(vals, n=3)  # [q1/3, q2/3]

    digest = hashlib.sha256(Path(xlsx_path).read_bytes()).hexdigest()
    return {
        "id": f"snow-t15-original-n{n}-seed{seed}-{REFERENCE_VERSION}",
        "built": _dt.date.today().isoformat(),
        "source": {
            "corpus": "SNOW T15 (Maruyama & Yamamoto, LREC 2018; CC BY 4.0)",
            "url": SNOW_URL,
            "file_sha256": digest,
            "column": "#日本語(原文)",
        },
        "sampling": f"uniform without replacement over {len(pairs)} pairs, "
        f"seed {seed}, n {n}; per-sentence signal values on originals",
        "n_used": {
            "jreadability": len(jread_vals),
            "lexical_coverage": len(cov_vals),
            "tmr": len(tmr_vals),
        },
        "signals": {
            "jreadability": {"edges": terciles(jread_vals), "higher_is_harder": False},
            "lexical_coverage": {"edges": terciles(cov_vals), "higher_is_harder": False},
            "tmr": {
                "edges": terciles(tmr_vals),
                "higher_is_harder": True,
                "level": core_level,
            },
        },
    }


def main(argv: list[str] | None = None) -> int:
    args = argv if argv is not None else sys.argv[1:]
    if not args:
        print("usage: python -m corpus.grading.reference <T15.xlsx> [n] [seed]", file=sys.stderr)
        return 2
    n = int(args[1]) if len(args) > 1 else 2000
    seed = int(args[2]) if len(args) > 2 else 42
    ref = build_reference(args[0], n=n, seed=seed)
    out = Path(__file__).resolve().parent / "reference_snow_t15.json"
    out.write_text(json.dumps(ref, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"reference {ref['id']} → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
